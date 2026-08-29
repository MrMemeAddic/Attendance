"""
tests_headless.py  -- Automated tests (no camera / no GUI required)
Run: .venv314\\Scripts\\python.exe tests_headless.py
"""
import sys, os, tempfile, pathlib
import numpy as np

os.environ.setdefault("PYTHONIOENCODING", "utf-8")
sys.stdout.reconfigure(encoding="utf-8")

_failures = []


def ok(label):
    print(f"  [PASS] {label}")


def fail(label, reason):
    print(f"  [FAIL] {label}: {reason}")
    _failures.append(label)


# -- Test 1: Database CRUD --------------------------------------------------
print("\n=== Test 1: Database CRUD ===")
try:
    from core.database import FaceDatabase

    tmp = pathlib.Path(tempfile.mkdtemp()) / "test.db"
    db = FaceDatabase(tmp)

    pid = db.add_person("Alice", "alice")
    assert pid == 1
    ok("add_person returns correct rowid")

    emb = np.random.randn(512).astype(np.float32)
    emb /= np.linalg.norm(emb)
    db.add_embedding(pid, emb)

    rows = db.get_all_embeddings()
    assert len(rows) == 1
    assert rows[0]["name"] == "Alice"
    assert rows[0]["embedding"].shape == (512,)
    ok("add_embedding + get_all_embeddings")

    cnt = db.count_embeddings(pid)
    assert cnt == 1
    ok("count_embeddings")

    db.log_recognition(pid, "Alice", 0.92, True)
    db.flush_log()                               # wait for async worker to commit
    log = db.get_recent_log(10)
    assert len(log) > 0, "Recognition log is empty after flush"
    assert log[0]["name"] == "Alice"
    assert abs(log[0]["confidence"] - 0.92) < 1e-5
    ok("log_recognition + get_recent_log")

    # Cascade delete: person gone => embeddings gone
    db.delete_person(pid)
    assert db.get_all_persons() == []
    assert db.get_all_embeddings() == []
    ok("delete_person cascade")

    pid2 = db.add_person("Bob", "bob")
    existing = db.get_person_by_label("bob")
    assert existing["id"] == pid2
    ok("get_person_by_label")

    db.close()
    ok("db.close")

except AssertionError as e:
    fail("Database CRUD", str(e))
except Exception as e:
    fail("Database CRUD", repr(e))


# -- Test 2: Matcher --------------------------------------------------------
print("\n=== Test 2: Matcher ===")
try:
    from core.matcher import FaceMatcher, cosine_similarity

    a = np.array([1.0, 0.0, 0.0], dtype=np.float32)
    b = np.array([1.0, 0.0, 0.0], dtype=np.float32)
    assert abs(cosine_similarity(a, b) - 1.0) < 1e-5
    ok("cosine_similarity identical vectors = 1.0")

    c = np.array([0.0, 1.0, 0.0], dtype=np.float32)
    assert abs(cosine_similarity(a, c) - 0.0) < 1e-5
    ok("cosine_similarity orthogonal vectors = 0.0")

    matcher = FaceMatcher(threshold=0.70)
    name, score = matcher.match(a)
    assert name == "Unknown"
    ok("empty gallery returns Unknown")

    matcher._gallery[1] = {"name": "Alice", "embeddings": [a.copy()]}
    matched_name, matched_score = matcher.match(a)
    assert matched_name == "Alice", f"got {matched_name}"
    assert abs(matched_score - 1.0) < 1e-4
    ok("exact match returns correct name + score=1.0")

    low = np.array([0.0, 0.0, 1.0], dtype=np.float32)
    name2, _ = matcher.match(low)
    assert name2 == "Unknown", f"expected Unknown, got {name2}"
    ok("below-threshold match returns Unknown")

    assert matcher.known_count == 1
    ok("known_count property")

except AssertionError as e:
    fail("Matcher", str(e))
except Exception as e:
    fail("Matcher", repr(e))


# -- Test 3: Liveness EAR ---------------------------------------------------
print("\n=== Test 3: Liveness / EAR ===")
try:
    from core.liveness import compute_ear, LivenessChecker, EAR_THRESHOLD

    open_pts   = [(0, 0), (1, 3), (3, 3), (6, 0), (3, -3), (1, -3)]
    closed_pts = [(0, 0), (1, 0), (3, 0), (6, 0), (3,  0), (1,  0)]

    ear_open = compute_ear(open_pts)
    assert ear_open > EAR_THRESHOLD, f"open EAR={ear_open:.3f} should be > {EAR_THRESHOLD}"
    ok(f"open EAR = {ear_open:.3f} > threshold ({EAR_THRESHOLD})")

    ear_closed = compute_ear(closed_pts)
    assert ear_closed < EAR_THRESHOLD, f"closed EAR={ear_closed:.3f} should be < {EAR_THRESHOLD}"
    ok(f"closed EAR = {ear_closed:.3f} < threshold ({EAR_THRESHOLD})")

    assert compute_ear([(0, 0)] * 3) == 1.0
    ok("short list returns 1.0 (open default)")

    # 3 closed frames then 1 open => 1 blink => PASSED (required=1)
    checker = LivenessChecker(required_blinks=1, timeout=10.0)
    checker.reset()
    for _ in range(3):
        checker.update(closed_pts, closed_pts)
    status = checker.update(open_pts, open_pts)
    assert checker.blink_count >= 1, f"blink_count={checker.blink_count}"
    assert status == "PASSED", f"expected PASSED, got {status}"
    ok("1-blink detection -> PASSED")

    # Timeout => FAILED
    import time
    checker2 = LivenessChecker(required_blinks=5, timeout=0.001)
    checker2.reset()
    time.sleep(0.05)
    status2 = checker2.update(open_pts, open_pts)
    assert status2 == "FAILED", f"expected FAILED, got {status2}"
    ok("timeout -> FAILED")

    status3 = checker2.update(closed_pts, closed_pts)
    assert status3 == "FAILED"
    ok("FAILED is terminal")

    # PASSED is terminal
    checker3 = LivenessChecker(required_blinks=1, timeout=10.0)
    checker3.reset()
    for _ in range(3):
        checker3.update(closed_pts, closed_pts)
    checker3.update(open_pts, open_pts)
    assert checker3.status == "PASSED"
    s = checker3.update(open_pts, open_pts)
    assert s == "PASSED"
    ok("PASSED is terminal")

except AssertionError as e:
    fail("Liveness/EAR", str(e))
except Exception as e:
    fail("Liveness/EAR", repr(e))


# -- Test 4: Tracker IoU ----------------------------------------------------
print("\n=== Test 4: Tracker _box_iou ===")
try:
    from core.tracker import _box_iou

    assert _box_iou((0, 0, 10, 10), (0, 0, 10, 10)) == 1.0
    ok("identical boxes -> IoU = 1.0")

    assert _box_iou((0, 0, 10, 10), (20, 20, 10, 10)) == 0.0
    ok("disjoint boxes -> IoU = 0.0")

    half = _box_iou((0, 0, 10, 10), (5, 0, 10, 10))
    assert abs(half - 1 / 3) < 0.01, f"expected ~0.333, got {half:.4f}"
    ok(f"50% overlap -> IoU ~ {half:.3f}")

    zero_union = _box_iou((0, 0, 0, 0), (0, 0, 0, 0))
    assert zero_union == 0.0
    ok("zero-area boxes -> IoU = 0.0 (no ZeroDivisionError)")

except AssertionError as e:
    fail("Tracker IoU", str(e))
except Exception as e:
    fail("Tracker IoU", repr(e))


# -- Test 5: Embedder _preprocess shape (no model load) ---------------------
print("\n=== Test 5: Embedder preprocess shape ===")
try:
    import torch, cv2
    from PIL import Image

    INPUT_SIZE = 160

    def _preprocess(face_bgr):
        rgb = cv2.cvtColor(face_bgr, cv2.COLOR_BGR2RGB)
        img = Image.fromarray(rgb).resize((INPUT_SIZE, INPUT_SIZE), Image.BILINEAR)
        arr = np.array(img, dtype=np.float32) / 255.0
        arr = (arr - 0.5) / 0.5
        tensor = torch.from_numpy(arr.transpose(2, 0, 1))
        return tensor.unsqueeze(0)

    dummy = np.zeros((100, 80, 3), dtype=np.uint8)
    t = _preprocess(dummy)
    assert t.shape == (1, 3, 160, 160), f"unexpected shape {t.shape}"
    ok("output shape is (1, 3, 160, 160)")

    assert float(t.min()) >= -1.0 and float(t.max()) <= 1.0
    ok("pixel values in [-1, 1]")

    dummy2 = np.random.randint(0, 255, (200, 60, 3), dtype=np.uint8)
    t2 = _preprocess(dummy2)
    assert t2.shape == (1, 3, 160, 160)
    ok("non-square crop also produces (1, 3, 160, 160)")

except AssertionError as e:
    fail("Embedder preprocess", str(e))
except Exception as e:
    fail("Embedder preprocess", repr(e))


# -- Test 6: Detector model paths -------------------------------------------
print("\n=== Test 6: Detector model paths ===")
try:
    from core.detector import _FD_MODEL_PATH, _FM_MODEL_PATH, _MODEL_DIR

    assert _MODEL_DIR.endswith("_models"), f"unexpected model dir: {_MODEL_DIR}"
    ok(f"_MODEL_DIR ends with _models")

    assert "core" in _FD_MODEL_PATH
    assert "core" in _FM_MODEL_PATH
    ok("model paths are under core/_models/")

except AssertionError as e:
    fail("Detector paths", str(e))
except Exception as e:
    fail("Detector paths", repr(e))


# -- Test 7: Tkinter availability (with TCL/TK path fix) --------------------
print("\n=== Test 7: Tkinter availability ===")
try:
    import sys, os as _os, sysconfig

    # Resolve the *base* Python installation (not the venv wrapper)
    # Try multiple candidate locations so the test passes both from run.bat
    # and from a plain PowerShell session.
    _candidates = [
        sysconfig.get_config_var("installed_base") or "",
        _os.path.dirname(sys.base_prefix),
        sys.base_prefix,
        getattr(sys, "base_prefix", ""),
        # Python.org Windows installer puts TCL here
        _os.path.join(_os.environ.get("LOCALAPPDATA", ""),
                      "Programs", "Python", "Python314"),
        _os.path.join(_os.environ.get("LOCALAPPDATA", ""),
                      "Programs", "Python", "Python313"),
        _os.path.join(_os.environ.get("LOCALAPPDATA", ""),
                      "Programs", "Python", "Python312"),
    ]
    for _base in _candidates:
        if not _base:
            continue
        _tcl_dir = _os.path.join(_base, "tcl")
        _tcl86   = _os.path.join(_tcl_dir, "tcl8.6")
        _tk86    = _os.path.join(_tcl_dir, "tk8.6")
        if _os.path.isfile(_os.path.join(_tcl86, "init.tcl")):
            _os.environ["TCL_LIBRARY"] = _tcl86
            _os.environ["TK_LIBRARY"]  = _tk86
            break

    # tkinter may already be partially imported; force a fresh Tk creation
    import tkinter as tk
    root = tk.Tk()
    root.withdraw()
    root.destroy()
    ok("Tkinter Tk() created and destroyed")
except Exception as e:
    fail("Tkinter", repr(e))
    print("  NOTE: run.bat sets TCL_LIBRARY/TK_LIBRARY before launching main.py,")
    print("        so the GUI will still work when double-clicking run.bat.")


# -- Test 8: AttendanceLogger -----------------------------------------------
print("\n=== Test 8: AttendanceLogger ===")
try:
    import time, pathlib, tempfile, csv
    import openpyxl
    from datetime import date as _date
    from core.attendance import AttendanceLogger

    att_dir = pathlib.Path(tempfile.mkdtemp()) / "attendance"
    logger  = AttendanceLogger(directory=att_dir, cooldown=2)

    # Log first entry
    result = logger.log("Alice", 0.88)
    assert result is True, "First log should succeed"
    ok("first log returns True")

    # Cooldown: same person immediately -> suppressed
    result2 = logger.log("Alice", 0.88)
    assert result2 is False, "Second log within cooldown should be False"
    ok("duplicate within cooldown returns False")

    # Unknown name -> always suppressed
    result3 = logger.log("Unknown", 0.50)
    assert result3 is False, "Unknown name should be suppressed"
    ok("'Unknown' name is suppressed")

    # Different person -> always logged
    result4 = logger.log("Bob", 0.75)
    assert result4 is True, "Different person should be logged"
    ok("different person logs immediately")

    # CSV file created
    csv_path = att_dir / f"attendance_{_date.today().isoformat()}.csv"
    assert csv_path.exists(), f"CSV not found at {csv_path}"
    ok("CSV file created")

    # CSV has correct columns and rows
    with open(csv_path, newline="", encoding="utf-8") as f:
        rows = list(csv.DictReader(f))
    assert len(rows) == 2, f"Expected 2 rows (Alice + Bob), got {len(rows)}"
    assert rows[0]["Name"] == "Alice"
    assert rows[1]["Name"] == "Bob"
    assert rows[0]["Status"] == "Present"
    ok(f"CSV has 2 rows with correct Name and Status columns")

    # XLSX file created and parseable
    xlsx_path = att_dir / f"attendance_{_date.today().isoformat()}.xlsx"
    assert xlsx_path.exists(), f"XLSX not found at {xlsx_path}"
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    assert ws.title == "Attendance"
    # Row 1 is header, rows 2+ are data
    headers = [ws.cell(1, c).value for c in range(1, 6)]
    assert headers == ["Date", "Time", "Name", "Confidence", "Status"], f"headers={headers}"
    data_rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert len(data_rows) == 2, f"Expected 2 data rows, got {len(data_rows)}"
    assert data_rows[0][2] == "Alice", f"row0 name={data_rows[0][2]}"
    assert data_rows[1][2] == "Bob",   f"row1 name={data_rows[1][2]}"
    assert data_rows[0][4] == "Present"
    # Check header cell is bold (styled)
    hdr_cell = ws.cell(1, 1)
    assert hdr_cell.font.bold, "Header cell should be bold"
    ok("XLSX file created with styled header and 2 data rows")

    # Cooldown expiry -> same person logged again
    time.sleep(2.1)
    result5 = logger.log("Alice", 0.91)
    assert result5 is True, "After cooldown, Alice should be logged again"
    ok("after cooldown expires, same person logged again")

    # get_today_records reflects all entries
    records_read = logger.get_records_for_date(_date.today())
    assert len(records_read) == 3, f"Expected 3 entries total, got {len(records_read)}"
    ok(f"get_records_for_date returns all {len(records_read)} entries")

    # list_attendance_dates includes today
    dates = logger.list_attendance_dates()
    assert _date.today() in dates
    ok("list_attendance_dates includes today")

except AssertionError as e:
    fail("AttendanceLogger", str(e))
except Exception as e:
    import traceback
    fail("AttendanceLogger", repr(e))
    traceback.print_exc()


# -- Test 9: Confidence gate (50 % floor) -----------------------------------
print("\n=== Test 9: Confidence gate (50 % floor) ===")
try:
    from core.matcher import FaceMatcher, UNKNOWN_LABEL
    from ui.recognize_panel import CONFIDENCE_THRESHOLD

    assert CONFIDENCE_THRESHOLD == 0.50, (
        f"Expected CONFIDENCE_THRESHOLD=0.50, got {CONFIDENCE_THRESHOLD}"
    )
    ok(f"CONFIDENCE_THRESHOLD is 0.50")

    # Build a tiny gallery with one known person
    a = np.array([1.0, 0.0, 0.0], dtype=np.float32)
    m = FaceMatcher(threshold=0.62)
    m._gallery[1] = {"name": "Alice", "embeddings": [a.copy()]}

    # Score >= 0.50 -> should keep the matched name
    high_query = np.array([1.0, 0.0, 0.0], dtype=np.float32)  # cosine = 1.0
    name_h, conf_h = m.match(high_query)
    # Apply gate logic (mirrors recognize_panel.py)
    if conf_h < CONFIDENCE_THRESHOLD:
        name_h = UNKNOWN_LABEL
    assert name_h == "Alice", f"Expected Alice, got {name_h} (conf={conf_h:.3f})"
    ok(f"conf={conf_h:.2f} >= 0.50 -> name kept as Alice")

    # Score < 0.50 -> gate must force Unknown
    low_query = np.array([0.0, 0.0, 1.0], dtype=np.float32)  # cosine = 0.0
    name_l, conf_l = m.match(low_query)
    if conf_l < CONFIDENCE_THRESHOLD:
        name_l = UNKNOWN_LABEL
    assert name_l == UNKNOWN_LABEL, f"Expected Unknown, got {name_l} (conf={conf_l:.3f})"
    ok(f"conf={conf_l:.2f} < 0.50 -> forced to Unknown")

    # Borderline: exactly 0.50 must NOT be blocked by the confidence gate.
    # Craft a vector whose cosine with a=[1,0,0] is exactly 0.5:
    #   cos(a, q) = q[0] / |q|  ->  q = [0.5, sqrt(0.75), 0]
    border_query = np.array([0.5, 0.8660254, 0.0], dtype=np.float32)
    _, conf_b = m.match(border_query)
    assert abs(conf_b - 0.5) < 0.01, f"Expected conf~0.50, got {conf_b:.4f}"
    # The gate uses strict '<', so conf==0.50 must NOT be blocked.
    gate_blocks = conf_b < CONFIDENCE_THRESHOLD
    assert not gate_blocks, (
        f"Gate incorrectly blocked conf={conf_b:.4f} which equals the threshold"
    )
    ok(f"conf={conf_b:.2f} at boundary: gate does NOT block (strict < threshold)")

except AssertionError as e:
    fail("Confidence gate", str(e))
except Exception as e:
    import traceback
    fail("Confidence gate", repr(e))
    traceback.print_exc()


# -- Summary ----------------------------------------------------------------
print("\n" + "=" * 54)
total = 9
passed = total - len(_failures)
print(f"RESULT: {passed}/{total} test groups passed")
if _failures:
    print(f"FAILED: {', '.join(_failures)}")
    sys.exit(1)
else:
    print("ALL TESTS PASSED")
    sys.exit(0)
